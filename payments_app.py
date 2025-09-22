import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import io
from datetime import datetime

# -----------------------------
# Helpers: lifetime & churn
# -----------------------------
def is_active_by_lifetime(age_from_activation:int,
                          age_from_first_payment:int,
                          lifetime_months:int,
                          lifetime_mode:str) -> bool:
    """ True nëse klienti është ende brenda dritares së lifetime. """
    if lifetime_months <= 0:
        return True
    if lifetime_mode == "From activation":
        return age_from_activation < lifetime_months
    else:
        return age_from_first_payment < lifetime_months

def survival_factor(churn_rate:float, churn_age:int) -> float:
    """ Faktor mbijetese me churn konstant mujor (0..1). """
    if churn_rate <= 0 or churn_age <= 0:
        return 1.0
    return (1.0 - churn_rate) ** churn_age

# -----------------------------
# Payment logic per contract
# -----------------------------
def per_client_payment(contract_type: str,
                       age_from_activation:int,
                       free_months:int,
                       intro_months:int,
                       intro_amount:float,
                       recurring_amount:float,
                       flat_amount:float) -> float:
    """ Pagesa bruto e NJË klienti në muajin e dhënë. """
    if age_from_activation < free_months:
        return 0.0
    eff_age = age_from_activation - free_months
    if contract_type == "Intro + Recurring":
        if eff_age < intro_months:
            return intro_amount
        return recurring_amount
    else:
        return flat_amount

# -----------------------------
# Core calculator
# -----------------------------
def build_monthly_dataframe(months:int,
                            default_new_clients:int,
                            override_month:int,              # 0 = off; ndryshe 1..months
                            override_new_clients:int,        # vlera për muajin e mbishkruar
                            cancellations_mode:str,          # "Fixed" or "Churn"
                            cancellations:int,               # vetëm kur Fixed
                            churn_percent:float,             # vetëm kur Churn
                            commission_rate:float,
                            payout_policy:str,               # "Bonus only", "Bonus + Recurring", "Recurring only"
                            payout_type:str,                 # "Commissionable (x%)", "Flat (direct)"
                            use_new_sale_bonus:bool,
                            new_sale_payout:float,           # shumë për New Client për muaj bonusi
                            payout_duration:int,             # në muaj nga fillimi i agjentit (jo nga klienti)
                            contract_type:str,               # "Intro + Recurring" | "Flat Monthly"
                            free_months:int,
                            intro_months:int,
                            intro_amount:float,
                            recurring_amount:float,
                            flat_amount:float,
                            lifetime_months:int,             # 0 = unlimited
                            lifetime_mode:str,               # "From activation" | "After free months"
                            currency:str) -> pd.DataFrame:

    churn_rate = max(min(churn_percent / 100.0, 0.99), 0.0)
    rows = []
    cohorts = []  # vetëm për recurring (jo për bonusin global)
    signed_cum = 0

    include_bonus = payout_policy in ["Bonus only (no recurring)", "Bonus + Recurring"]
    include_recurring = payout_policy in ["Bonus + Recurring", "Recurring only (no bonus)"]
    bonus_is_commissionable = (payout_type == "Commissionable (x%)")

    for m in range(1, months + 1):
        # NEW clients për këtë muaj (me override nëse aplikohet)
        new_clients_this_month = override_new_clients if (override_month == m and override_month > 0) else default_new_clients

        # Cohort për recurring
        if cancellations_mode == "Fixed":
            net_size = max(new_clients_this_month - cancellations, 0)
        else:
            # Churn mode: nuk zbriten cancellations në hyrje; i aplikohet churn më vonë
            net_size = new_clients_this_month
        cohorts.append({"birth": m, "size": net_size})

        signed_cum += net_size

        gross_client_payments = 0.0
        paying_clients_this_month = 0.0

        # Recurring nga të gjithë cohort-et
        if include_recurring:
            for c in cohorts:
                age_from_activation = m - c["birth"]
                if age_from_activation < 0:
                    continue
                age_from_first_payment = max(age_from_activation - free_months, 0)

                # Lifetime gating
                active_by_life = is_active_by_lifetime(
                    age_from_activation=age_from_activation,
                    age_from_first_payment=age_from_first_payment,
                    lifetime_months=lifetime_months,
                    lifetime_mode=lifetime_mode
                )
                if not active_by_life:
                    continue

                # Churn (vetëm në "Churn mode")
                if cancellations_mode == "Churn":
                    churn_age = age_from_activation if lifetime_mode == "From activation" else age_from_first_payment
                    sf = survival_factor(churn_rate, churn_age)
                else:
                    sf = 1.0  # në "Fixed" e kemi reflektuar në net_size

                per_client_gross = per_client_payment(
                    contract_type=contract_type,
                    age_from_activation=age_from_activation,
                    free_months=free_months,
                    intro_months=intro_months,
                    intro_amount=intro_amount,
                    recurring_amount=recurring_amount,
                    flat_amount=flat_amount
                )
                if per_client_gross > 0:
                    active_clients = c["size"] * sf
                    paying_clients_this_month += active_clients
                    gross_client_payments += per_client_gross * active_clients

        commission_from_clients = gross_client_payments * commission_rate if include_recurring else 0.0

        # Bonus global sipas muajve të parë të AGJENTIT (m <= payout_duration)
        bonus_raw = 0.0
        if include_bonus and use_new_sale_bonus and payout_duration > 0 and m <= payout_duration:
            bonus_raw = new_sale_payout * new_clients_this_month  # BAZUAR te NEW clients të këtij muaji

        bonus_to_agent = (bonus_raw * commission_rate) if bonus_is_commissionable else bonus_raw

        # Total
        if payout_policy == "Bonus only (no recurring)":
            total = bonus_to_agent
        elif payout_policy == "Recurring only (no bonus)":
            total = commission_from_clients
        else:
            total = bonus_to_agent + commission_from_clients

        rows.append({
            "Month": m,
            "New Clients": new_clients_this_month,
            "Cancellations": cancellations if cancellations_mode == "Fixed" else "-",
            "Mode": cancellations_mode,
            "Net Activations": net_size,
            "Signed SMEs (Cumulative)": signed_cum,
            "Paying SMEs (this month)": paying_clients_this_month,
            "Client Payments (Gross)": gross_client_payments,
            "New Sale Income": bonus_to_agent,
            "Commission from Clients": commission_from_clients,
            "Total Monthly Earnings": total
        })

    return pd.DataFrame(rows)

def yearly_totals(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out["Year"] = (out["Month"] - 1) // 12 + 1
    out = out.groupby("Year", as_index=False).agg({
        "Client Payments (Gross)": "sum",
        "New Sale Income": "sum",
        "Commission from Clients": "sum",
        "Total Monthly Earnings": "sum"
    })
    out = out.rename(columns={"Total Monthly Earnings": "Total Yearly Earnings"})
    return out

def to_excel_values(df: pd.DataFrame, inputs: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws["A1"] = "Inputs"
    r = 2
    for k, v in inputs.items():
        ws[f"A{r}"] = k
        ws[f"B{r}"] = v
        r += 1

    headers = list(df.columns)
    start_row = r + 1
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=col_idx).value = h

    for i, (_, row) in enumerate(df.iterrows(), start=1):
        for j, h in enumerate(headers, start=1):
            val = row[h]
            ws.cell(row=start_row + i, column=j).value = float(val) if isinstance(val, (int, float)) else val

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 24

    ws2 = wb.create_sheet("Foglio1")
    ws2["A1"] = "Payments – Summary"
    ws2["A2"] = "Generated"
    ws2["B2"] = datetime.now().strftime("%Y-%m-%d %H:%M")

    yt = yearly_totals(df)
    if not yt.empty:
        ws2["A4"] = "Year"
        ws2["B4"] = "Client Payments (Gross)"
        ws2["C4"] = "New Sale Income"
        ws2["D4"] = "Commission from Clients"
        ws2["E4"] = "Total Yearly Earnings"

        for idx, row in yt.iterrows():
            base = 5 + idx
            ws2[f"A{base}"] = int(row["Year"])
            ws2[f"B{base}"] = float(row["Client Payments (Gross)"])
            ws2[f"C{base}"] = float(row["New Sale Income"])
            ws2[f"D{base}"] = float(row["Commission from Clients"])
            ws2[f"E{base}"] = float(row["Total Yearly Earnings"])

    for col in range(1, 7):
        ws2.column_dimensions[get_column_letter(col)].width = 28

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()

# -----------------------------
# UI
# -----------------------------
st.set_page_config(page_title="Earnings Calculator (Contracts v3.3)",
                   page_icon="💼",
                   layout="wide")

TABLE_HEIGHT = 480

st.title("Movylo – Earnings Calculator (Contracts v3.3)")
st.caption("Shto test override: cakto New Clients vetëm për një muaj të zgjedhur. Mbështetur Fixed/Churn, Lifetime, Bonus duration, Export.")

with st.sidebar:
    st.header("Scope & Horizon")
    months = st.slider("Months to project", min_value=12, max_value=60, value=36, step=12)
    currency = st.text_input("Currency symbol", value="£")

    st.header("Funnel")
    default_new_clients = st.number_input("New Clients per Month (default)", min_value=0, value=12, step=1)

    # TEST OVERRIDE – vendos New Clients vetëm në një muaj specifik
    st.markdown("**Test month override**")
    use_override = st.checkbox("Override a specific month", value=False)
    if use_override:
        override_month = st.number_input("Month to override (1..projection)", min_value=1, max_value=months, value=5, step=1)
        override_new_clients = st.number_input("New Clients for that month", min_value=0, value=20, step=1)
    else:
        override_month = 0
        override_new_clients = 0

    cancellations_mode = st.radio(
        "Cancellations mode",
        options=["Fixed cancellations per month", "Churn % per active month"],
        index=0
    )
    if cancellations_mode.startswith("Fixed"):
        cancellations = st.number_input("Cancellations per Month", min_value=0, value=2, step=1)
        churn_percent = 0.0
        mode_key = "Fixed"
    else:
        cancellations = 0  # nuk përdoret
        churn_percent = st.slider("Churn % per active month", min_value=0, max_value=50, value=0, step=1)
        mode_key = "Churn"

    st.header("Lifetime")
    lifetime_months = st.number_input("Average client lifetime (months)", min_value=0, value=0, step=1,
                                      help="0 = unlimited")
    lifetime_mode = st.radio("Lifetime counting mode",
                             options=["From activation", "After free months"], index=0)

    st.header("Contract Plan")
    contract_type = st.radio("Type", options=["Intro + Recurring", "Flat Monthly"])
    free_months = st.number_input("Free Months at start", min_value=0, max_value=24, value=0, step=1)

    if contract_type == "Intro + Recurring":
        intro_months = st.number_input("Intro Months (minimum term)", min_value=1, max_value=24, value=3, step=1)
        intro_amount = st.number_input(f"Intro Monthly Amount ({currency})", min_value=0.0, value=300.0, step=10.0)
        recurring_amount = st.number_input(f"Recurring Monthly Amount ({currency})", min_value=0.0, value=150.0, step=10.0)
        flat_amount = 0.0
    else:
        intro_months = 0
        intro_amount = 0.0
        recurring_amount = 0.0
        flat_amount = st.number_input(f"Flat Monthly Amount ({currency})", min_value=0.0, value=100.0, step=10.0)

    st.header("Commission & Bonus")
    commission_pct = st.slider("Commission Rate (%)", min_value=0, max_value=100, value=80, step=5)
    commission_rate = commission_pct / 100.0

    payout_policy = st.selectbox("Payout policy",
                                 options=["Bonus only (no recurring)",
                                          "Bonus + Recurring",
                                          "Recurring only (no bonus)"], index=1)
    payout_type = st.radio("Payout type",
                           options=["Commissionable (x%)", "Flat (direct)"], index=0)

    use_new_sale_bonus = st.checkbox("Use one-off New Sale Payout", value=True)
    new_sale_payout = st.number_input("New Sale Payout per New Client", min_value=0.0, value=160.0, step=10.0,
                                      disabled=not use_new_sale_bonus)
    payout_duration = st.number_input(
        "Payout duration (months)",
        min_value=0, max_value=24, value=1, step=1,
        help="0 = off. Paguhet vetëm muajt e parë TË AGJENTIT (p.sh. janar–shkurt), pastaj 0; i pavarur nga zgjatja e klientit."
    )

df = build_monthly_dataframe(
    months=months,
    default_new_clients=default_new_clients,
    override_month=override_month,
    override_new_clients=override_new_clients,
    cancellations_mode=mode_key,
    cancellations=cancellations,
    churn_percent=churn_percent,
    commission_rate=commission_rate,
    payout_policy=payout_policy,
    payout_type=payout_type,
    use_new_sale_bonus=use_new_sale_bonus,
    new_sale_payout=new_sale_payout,
    payout_duration=payout_duration,
    contract_type=contract_type,
    free_months=free_months,
    intro_months=intro_months,
    intro_amount=intro_amount,
    recurring_amount=recurring_amount,
    flat_amount=flat_amount,
    lifetime_months=lifetime_months,
    lifetime_mode=lifetime_mode,
    currency=currency
)

# Shfaqja e tabelave
col1, col2 = st.columns([2, 1])
with col1:
    st.subheader("Monthly Breakdown")
    st.dataframe(
        df.style.format({
            "Client Payments (Gross)": f"{currency}" + "{:,.2f}",
            "New Sale Income": f"{currency}" + "{:,.2f}",
            "Commission from Clients": f"{currency}" + "{:,.2f}",
            "Total Monthly Earnings": f"{currency}" + "{:,.2f}",
        }),
        use_container_width=True,
        height=TABLE_HEIGHT
    )

with col2:
    yt = yearly_totals(df)
    st.subheader("Yearly Totals")
    st.dataframe(
        yt.style.format({
            "Client Payments (Gross)": f"{currency}" + "{:,.2f}",
            "New Sale Income": f"{currency}" + "{:,.2f}",
            "Commission from Clients": f"{currency}" + "{:,.2f}",
            "Total Yearly Earnings": f"{currency}" + "{:,.2f}",
        }),
        use_container_width=True,
        height=TABLE_HEIGHT
    )

st.subheader("Trend")
trend_df = df[["Month", "Client Payments (Gross)", "Commission from Clients", "Total Monthly Earnings"]].set_index("Month")
st.line_chart(trend_df)

# Export
st.divider()
st.subheader("Export")
excel_bytes = to_excel_values(
    df=df,
    inputs={
        "Months": months,
        "Currency": currency,
        "New Clients / month (default)": default_new_clients,
        "Override month": override_month if override_month > 0 else "-",
        "Override New Clients": override_new_clients if override_month > 0 else "-",
        "Cancellations mode": mode_key,
        "Cancellations / month (if Fixed)": cancellations if mode_key=="Fixed" else "-",
        "Churn % (if Churn)": churn_percent if mode_key=="Churn" else "-",
        "Lifetime (months)": lifetime_months,
        "Lifetime mode": lifetime_mode,
        "Contract Type": contract_type,
        "Free Months": free_months,
        "Intro Months": intro_months,
        "Intro Amount": intro_amount,
        "Recurring Amount": recurring_amount,
        "Flat Amount": flat_amount,
        "Commission Rate (%)": commission_pct,
        "Payout policy": payout_policy,
        "Payout type": payout_type,
        "Use New Sale Payout": use_new_sale_bonus,
        "New Sale Payout": new_sale_payout,
        "Payout duration (months)": payout_duration
    }
)
st.download_button(
    label="Download as Excel",
    data=excel_bytes,
    file_name="Payments_v3_3.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)